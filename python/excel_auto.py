from openpyxl import xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell = sheet['a1'] # row col according to given in excel
cell = sheet.cell(1,1) #row,col
for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    correct = cell.value * 0.9
    sheet.cell(row,4).value = correct


values = Reference(sheet,
                   min_row = 2,
                   max_row = sheet.max_row,
                   min_col= 4,
                   max_col =4
                        )
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,"e2")
wb.save("trans.xlsx")     